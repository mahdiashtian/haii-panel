from datetime import datetime
import os

import openpyxl
from django.utils.crypto import get_random_string
from persiantools.jdatetime import JalaliDate


def convertor_hijri(year, mounth, day):
    # convert Gregorian date to Solar Hijri date tuple

    # create a Jalali (Persian) date object from the Gregorian date
    persian_date = JalaliDate.to_jalali(int(year), int(mounth), int(day))

    # format the Persian date object as a string in the format yyyy-mm-dd
    persian_string = persian_date.strftime('%Y-%m-%d')

    return persian_string


def get_file_name(file_name):
    """
    _Get File Name_
    Args:
        file_name (_str_): Name file (image,movie and ...)
    Returns:
        _tuple_: Name and ext
    """

    base_name = os.path.basename(file_name)
    name, ext = os.path.splitext(base_name)
    return name, ext


def upload_image_path(instance=None, filename=None):
    """
    _Upload Image Path_
    Args:
        instance (django obj instance, optional): The object that the file is for. Defaults to None.
        filename (_type_, optional): The final generated name for the file. Defaults to None.
    Returns:
        _str_: Final path and file name
    """

    model_name = instance.__class__.__name__.lower()
    name, ext = get_file_name(filename)
    return f"{model_name}/{get_random_string(10)}/{name}{ext}"


def get_country_list():
    """
    _Get Country List_
    Returns:
        _list_: List of countries
    """

    return (
        ('AF', 'Afghanistan'), ('AX', 'Aland Islands'), ('AL', 'Albania'), ('DZ', 'Algeria'), ('AS', 'American Samoa'),
        ('AD', 'Andorra'), ('AO', 'Angola'), ('AI', 'Anguilla'), ('AQ', 'Antarctica'), ('AG', 'Antigua & Barbuda'),
        ('AR', 'Argentina'), ('AM', 'Armenia'), ('AW', 'Aruba'), ('AU', 'Australia'), ('AT', 'Austria'),
        ('AZ', 'Azerbaijan'), ('BS', 'Bahamas'), ('BH', 'Bahrain'), ('BD', 'Bangladesh'), ('BB', 'Barbados'),
        ('BY', 'Belarus'), ('BE', 'Belgium'), ('BZ', 'Belize'), ('BJ', 'Benin'), ('BM', 'Bermuda'), ('BT', 'Bhutan'),
        ('BO', 'Bolivia'), ('BA', 'Bosnia & Herzegovina'), ('BW', 'Botswana'), ('BV', 'Bouvet Island'),
        ('BR', 'Brazil'),
        ('IO', 'British Indian Ocean Territory'), ('VG', 'British Virgin Islands'), ('BN', 'Brunei'),
        ('BG', 'Bulgaria'),
        ('BF', 'Burkina Faso'), ('BI', 'Burundi'), ('KH', 'Cambodia'), ('CM', 'Cameroon'), ('CA', 'Canada'),
        ('CV', 'Cape Verde'), ('BQ', 'Caribbean Netherlands'), ('KY', 'Cayman Islands'),
        ('CF', 'Central African Republic'),
        ('TD', 'Chad'), ('CL', 'Chile'), ('CN', 'China'), ('CX', 'Christmas Island'), ('CC', 'Cocos (Keeling) Islands'),
        ('CO', 'Colombia'), ('KM', 'Comoros'), ('CG', 'Congo - Brazzaville'), ('CD', 'Congo - Kinshasa'),
        ('CK', 'Cook Islands'), ('CR', 'Costa Rica'), ('CI', 'Côte d’Ivoire'), ('HR', 'Croatia'), ('CU', 'Cuba'),
        ('CW', 'Curaçao'), ('CY', 'Cyprus'), ('CZ', 'Czechia'), ('DK', 'Denmark'), ('DJ', 'Djibouti'),
        ('DM', 'Dominica'),
        ('DO', 'Dominican Republic'), ('EC', 'Ecuador'), ('EG', 'Egypt'), ('SV', 'El Salvador'),
        ('GQ', 'Equatorial Guinea'), ('ER', 'Eritrea'), ('EE', 'Estonia'), ('SZ', 'Eswatini'), ('ET', 'Ethiopia'),
        ('FK', 'Falkland Islands'), ('FO', 'Faroe Islands'), ('FJ', 'Fiji'), ('FI', 'Finland'), ('FR', 'France'),
        ('GF', 'French Guiana'), ('PF', 'French Polynesia'), ('TF', 'French Southern Territories'), ('GA', 'Gabon'),
        ('GM', 'Gambia'), ('GE', 'Georgia'), ('DE', 'Germany'), ('GH', 'Ghana'), ('GI', 'Gibraltar'), ('GR', 'Greece'),
        ('GL', 'Greenland'), ('GD', 'Grenada'), ('GP', 'Guadeloupe'), ('GU', 'Guam'), ('GT', 'Guatemala'),
        ('GG', 'Guernsey'), ('GN', 'Guinea'), ('GW', 'Guinea-Bissau'), ('GY', 'Guyana'), ('HT', 'Haiti'),
        ('HM', 'Heard & McDonald Islands'), ('HN', 'Honduras'), ('HK', 'Hong Kong SAR China'), ('HU', 'Hungary'),
        ('IS', 'Iceland'), ('IN', 'India'), ('ID', 'Indonesia'), ('IR', 'Iran'), ('IQ', 'Iraq'), ('IE', 'Ireland'),
        ('IM', 'Isle of Man'), ('IL', 'Israel'), ('IT', 'Italy'), ('JM', 'Jamaica'), ('JP', 'Japan'), ('JE', 'Jersey'),
        ('JO', 'Jordan'), ('KZ', 'Kazakhstan'), ('KE', 'Kenya'), ('KI', 'Kiribati'), ('KW', 'Kuwait'),
        ('KG', 'Kyrgyzstan'),
        ('LA', 'Laos'), ('LV', 'Latvia'), ('LB', 'Lebanon'), ('LS', 'Lesotho'), ('LR', 'Liberia'), ('LY', 'Libya'),
        ('LI', 'Liechtenstein'), ('LT', 'Lithuania'), ('LU', 'Luxembourg'), ('MO', 'Macao SAR China'),
        ('MG', 'Madagascar'),
        ('MW', 'Malawi'), ('MY', 'Malaysia'), ('MV', 'Maldives'), ('ML', 'Mali'), ('MT', 'Malta'),
        ('MH', 'Marshall Islands'), ('MQ', 'Martinique'), ('MR', 'Mauritania'), ('MU', 'Mauritius'), ('YT', 'Mayotte'),
        ('MX', 'Mexico'), ('FM', 'Micronesia'), ('MD', 'Moldova'), ('MC', 'Monaco'), ('MN', 'Mongolia'),
        ('ME', 'Montenegro'), ('MS', 'Montserrat'), ('MA', 'Morocco'), ('MZ', 'Mozambique'), ('MM', 'Myanmar (Burma)'),
        ('NA', 'Namibia'), ('NR', 'Nauru'), ('NP', 'Nepal'), ('NL', 'Netherlands'), ('NC', 'New Caledonia'),
        ('NZ', 'New Zealand'), ('NI', 'Nicaragua'), ('NE', 'Niger'), ('NG', 'Nigeria'), ('NU', 'Niue'),
        ('NF', 'Norfolk Island'), ('KP', 'North Korea'), ('MK', 'North Macedonia'), ('MP', 'Northern Mariana Islands'),
        ('NO', 'Norway'), ('OM', 'Oman'), ('PK', 'Pakistan'), ('PW', 'Palau'), ('PS', 'Palestinian Territories'),
        ('PA', 'Panama'), ('PG', 'Papua New Guinea'), ('PY', 'Paraguay'), ('PE', 'Peru'), ('PH', 'Philippines'),
        ('PN', 'Pitcairn Islands'), ('PL', 'Poland'), ('PT', 'Portugal'), ('PR', 'Puerto Rico'), ('QA', 'Qatar'),
        ('RE', 'Réunion'), ('RO', 'Romania'), ('RU', 'Russia'), ('RW', 'Rwanda'), ('WS', 'Samoa'), ('SM', 'San Marino'),
        ('ST', 'São Tomé & Príncipe'), ('SA', 'Saudi Arabia'), ('SN', 'Senegal'), ('RS', 'Serbia'),
        ('SC', 'Seychelles'),
        ('SL', 'Sierra Leone'), ('SG', 'Singapore'), ('SX', 'Sint Maarten'), ('SK', 'Slovakia'), ('SI', 'Slovenia'),
        ('SB', 'Solomon Islands'), ('SO', 'Somalia'), ('ZA', 'South Africa'),
        ('GS', 'South Georgia & South Sandwich Islands'), ('KR', 'South Korea'), ('SS', 'South Sudan'), ('ES', 'Spain'),
        ('LK', 'Sri Lanka'), ('BL', 'St. Barthélemy'), ('SH', 'St. Helena'), ('KN', 'St. Kitts & Nevis'),
        ('LC', 'St. Lucia'), ('MF', 'St. Martin'), ('PM', 'St. Pierre & Miquelon'), ('VC', 'St. Vincent & Grenadines'),
        ('SD', 'Sudan'), ('SR', 'Suriname'), ('SJ', 'Svalbard & Jan Mayen'), ('SE', 'Sweden'), ('CH', 'Switzerland'),
        ('SY', 'Syria'), ('TW', 'Taiwan'), ('TJ', 'Tajikistan'), ('TZ', 'Tanzania'), ('TH', 'Thailand'),
        ('TL', 'Timor-Leste'), ('TG', 'Togo'), ('TK', 'Tokelau'), ('TO', 'Tonga'), ('TT', 'Trinidad & Tobago'),
        ('TN', 'Tunisia'), ('TR', 'Turkey'), ('TM', 'Turkmenistan'), ('TC', 'Turks & Caicos Islands'), ('TV', 'Tuvalu'),
        ('UM', 'U.S. Outlying Islands'), ('VI', 'U.S. Virgin Islands'), ('UG', 'Uganda'), ('UA', 'Ukraine'),
        ('AE', 'United Arab Emirates'), ('GB', 'United Kingdom'), ('US', 'United States'), ('UY', 'Uruguay'),
        ('UZ', 'Uzbekistan'), ('VU', 'Vanuatu'), ('VA', 'Vatican City'), ('VE', 'Venezuela'), ('VN', 'Vietnam'),
        ('WF', 'Wallis & Futuna'), ('EH', 'Western Sahara'), ('YE', 'Yemen'), ('ZM', 'Zambia'), ('ZW', 'Zimbabwe'))


def convert_to_excel(data):
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Define headers and subheaders
    headers = [
        ('اطلاعات شخصی ', ['نام', 'نام خانوادگی', 'تاریخ ولد', 'کد ملی',
                           'ملیت', 'استان', 'شهر', 'آدرس', ',وضعیت تاهل', 'تعداد فرزندان']),
        ('اطلاعات تحصیلی', ['رشته تحصیلی', 'مقطع تحصیلی', 'محل تحصیل']),
        ('اطلاعات تیمی', ['نام تیم', 'بخش های فعالیت', 'مدیر تیم']),
        ('تجربه کاری', ['عنوان', 'نام کمپانی', 'مدت زمان'])
    ]

    # Keep track of the current row and column
    row = 1
    col = 1

    # Add headers and subheaders to the sheet
    for header, subheaders in headers:
        # Add the header to the sheet
        sheet.cell(row=row, column=col, value=header)
        sheet.cell(row=row, column=col, value=header)
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(subheaders) - 1)

        # Add the subheaders to the sheet
        for i, subheader in enumerate(subheaders):
            sheet.cell(row=row + 1, column=col + i, value=subheader)
            sheet.cell(row=row + 1, column=col + i, value=subheader)

        # Update the current row and column
        # row = 1
        col += len(subheaders)

    # Add data to the sheet
    row += 2

    for i, item in enumerate(data):
        sheet.cell(row=i + row, column=1, value=item['first_name'])
        sheet.cell(row=i + row, column=2, value=item['last_name'])
        date = item['date_of_birth'].split("-")
        date_of_birth = convertor_hijri(date[0], date[1], date[2])

        sheet.cell(row=i + row, column=3, value=date_of_birth)

        if item.get('iranian_profile', False):
            sheet.cell(row=i + row, column=4, value=item['iranian_profile']['national_code'])
        else:
            sheet.cell(row=i + row, column=4, value=item['foreigner_profile']['exclusive_code'])

        sheet.cell(row=i + row, column=5, value=item['country'])
        sheet.cell(row=i + row, column=6, value=item['state'])
        sheet.cell(row=i + row, column=7, value=item['city'])
        sheet.cell(row=i + row, column=8, value=item['address'])
        if item['marital_status'] == 'married':
            item['marital_status'] = 'متاهل'
        else:
            item['marital_status'] = 'مجرد'
        sheet.cell(row=i + row, column=9, value=item['marital_status'])
        if item.get("child"):
            sheet.cell(row=i + row, column=10, value=len(item['child']))

        # set last education
        item['education_profile'] = item['education_profile'][-1]

        # create match and case for education
        if item['education_profile']['grade'] == 'CY':
            item['education_profile']['grade'] = 'سیکل'
        elif item['education_profile']['grade'] == 'DI':
            item['education_profile']['grade'] = 'دیپلم'
        elif item['education_profile']['grade'] == 'MA':
            item['education_profile']['grade'] = 'کارشناسی'
        elif item['education_profile']['grade'] == 'MP':
            item['education_profile']['grade'] = 'کارشناسی ارشد'
        elif item['education_profile']['grade'] == 'DA':
            item['education_profile']['grade'] = 'دکتری'

        sheet.cell(row=i + row, column=11, value=item['education_profile']['major'])
        sheet.cell(row=i + row, column=12, value=item['education_profile']['grade'])
        sheet.cell(row=i + row, column=13, value=item['education_profile']['name'])
        #
        # sheet.cell(row=i + row, column=14, value=item['team_name'])
        # sheet.cell(row=i + row, column=15, value=item['activity_sections'])
        # sheet.cell(row=i + row, column=6, value=item['team_manager'])
        #

        name = [i['name'] for i in item['experience_profile']]
        company = [i['company'] for i in item['experience_profile']]

        time = [str((datetime.strptime(i['stop'], '%Y-%m-%d') - datetime.strptime(i['start'], '%Y-%m-%d')).days) for i in
                item['experience_profile']]
        time = []

        for work in item['experience_profile']:
            time_work = (datetime.strptime(work['stop'], '%Y-%m-%d') - datetime.strptime(work['start'], '%Y-%m-%d'))
            days = time_work.days
            years, days = divmod(days, 365)
            months, days = divmod(days, 30)
            time.append(f"{years} سال و {months} ماه و {days} روز")

        sheet.cell(row=i + row, column=17, value="\n".join(name))
        sheet.cell(row=i + row, column=18, value="\n".join(company))
        sheet.cell(row=i + row, column=19, value="\n".join(time))

    # Save the workbook
    path = os.path.dirname(os.path.abspath(__file__))
    path_file = f'{path}/data.xlsx'

    sheet.sheet_view.rightToLeft = True
    if os.path.exists(path_file):
        os.remove(path_file)
    workbook.save(path_file)
    return path_file
